"""Extract CA current features for all 115 groups, to build the CA-side disconnect classifier.
Reads each group's CA data.csv (# potential_V=x header, then time,current), computes the
current-scale / saturation / fit-quality features that fingerprint electrode disconnection,
and joins Book1 ground-truth labels. Output: ca_disconnect_feats.csv for threshold-setting.

Usage:
    python ca_disconnect_features.py <ca_root> [book1.xlsx]
      <ca_root>  directory holding per-group CA data, as <ca_root>/<group>/data.csv
                 (group = 1..115). If your layout differs, edit ca_path().
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import numpy as np

try:
    import openpyxl
except ImportError:
    openpyxl = None


def ca_path(root: Path, g: int) -> Path:
    """Where group g's CA data.csv lives. EDIT THIS if your layout differs."""
    return root / str(g) / "data.csv"


def load_ca(path: Path):
    """CA data.csv -> (potential_V, t, i). Header '# potential_V=x', then 'time,current'."""
    pot = np.nan
    ts, cs = [], []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("#"):
                if "potential_V=" in line:
                    pot = float(line.split("potential_V=")[1])
                continue
            if line.lower().startswith("time"):
                continue
            parts = line.split(",")
            if len(parts) < 2 or parts[1] == "":
                continue
            ts.append(float(parts[0])); cs.append(float(parts[1]))
    return pot, np.array(ts), np.array(cs)


def succ_diff_mad(x):
    """Robust noise: MAD of successive differences / (1.4826*sqrt(2)); immune to slow trend."""
    if len(x) < 3:
        return np.nan
    d = np.diff(x)
    return np.median(np.abs(d - np.median(d))) * 1.4826 / np.sqrt(2)


def ca_features(t, i):
    """Current-domain fingerprint features (no model fit needed for the connection gate).
    max_abs_i     : peak |current| -- primary scale discriminator (open -> tiny, ref-loss -> huge)
    tail_mean     : mean current over last 20% -- steady-state level
    first_i       : |current| of first sample -- initial transient size
    sigma_noise   : successive-difference MAD -- noise floor
    saturation_frac: fraction of |i| within 1% of the run's max |i| -- ADC railing (working-loss high)
    sign_flips    : fraction of adjacent samples that change sign -- pure-noise traces flip a lot
    """
    if len(i) < 5:
        return None
    abs_i = np.abs(i)
    mx = float(abs_i.max())
    n_tail = max(1, len(i) // 5)
    sat = float(np.mean(abs_i >= 0.99 * mx)) if mx > 0 else np.nan
    flips = float(np.mean(np.diff(np.sign(i)) != 0)) if len(i) > 1 else np.nan
    return {
        "max_abs_i": mx,
        "tail_mean": float(np.mean(i[-n_tail:])),
        "first_i": float(abs_i[0]),
        "sigma_noise": succ_diff_mad(i),
        "saturation_frac": sat,
        "sign_flips": flips,
        "n_points": len(i),
    }


def labels_from_book1(xlsx: Path):
    """group -> (substance, disconnect, polished, potential, day) from Book1 Sheet1."""
    if openpyxl is None:
        return {}
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Sheet1"]
    out = {}
    disc_map = {"X": "none", "working": "working", "counter": "counter", "reference": "reference"}
    for r in range(2, ws.max_row + 1):
        g = ws.cell(r, 1).value
        if g is None:
            continue
        out[int(g)] = {
            "substance": ws.cell(r, 2).value,
            "potential": ws.cell(r, 3).value,
            "polished": ws.cell(r, 4).value == "有",
            "disconnect": disc_map.get(str(ws.cell(r, 5).value).strip(), str(ws.cell(r, 5).value).strip()),
            "day": ws.cell(r, 6).value,
        }
    return out


def main(ca_root, xlsx="Book1.xlsx"):
    root = Path(ca_root)
    labels = labels_from_book1(Path(xlsx))
    rows = []
    missing = []
    for g in range(1, 116):
        p = ca_path(root, g)
        if not p.is_file():
            missing.append(g)
            continue
        try:
            pot, t, i = load_ca(p)
            feat = ca_features(t, i)
            if feat is None:
                missing.append(g); continue
            row = {"group": g, "potential_file": pot, **feat}
            row.update(labels.get(g, {}))
            rows.append(row)
        except Exception as e:
            print(f"  group {g}: {e}")
            missing.append(g)

    if not rows:
        print(f"NO CA data found under {root}. Check ca_path() layout.")
        print(f"tried e.g. {ca_path(root,36)}")
        return
    out = Path("ca_disconnect_feats.csv")
    fields = ["group", "substance", "disconnect", "polished", "potential", "potential_file",
              "day", "max_abs_i", "tail_mean", "first_i", "sigma_noise",
              "saturation_frac", "sign_flips", "n_points"]
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)
    print(f"wrote {out}: {len(rows)} groups")
    if missing:
        print(f"missing/unreadable: {len(missing)} groups -> {missing}")

    # quick per-class summary so you (and I) can eyeball separability immediately
    import collections
    by = collections.defaultdict(list)
    for r in rows:
        by[r.get("disconnect", "?")].append(r)
    print("\n=== CA features by disconnect (median) ===")
    for dc in ["none", "working", "counter", "reference"]:
        s = by.get(dc, [])
        if not s:
            continue
        def med(k): return np.nanmedian([r[k] for r in s])
        print(f"  {dc:10} n={len(s):3d}  max_abs_i={med('max_abs_i'):.2e}  "
              f"sat={med('saturation_frac'):.3f}  sigma={med('sigma_noise'):.2e}  "
              f"flips={med('sign_flips'):.3f}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python ca_disconnect_features.py <ca_root> [book1.xlsx]")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "Book1.xlsx")